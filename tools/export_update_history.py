from __future__ import annotations

import csv
import html
import subprocess
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
CSV_PATH = ROOT / "tabiroku-update-history.csv"
XLSX_PATH = ROOT / "tabiroku-update-history.xlsx"
HTML_PATH = ROOT / "update-history.html"
PUBLIC_URL = "https://ryoseiimai.github.io/tabiroku-app/"
REPO_URL = "https://github.com/Ryoseiimai/tabiroku-app"

DETAIL_MAP = {
    "Initial Tabiroku mobile mock": {
        "detail": "旅録モックの初期画面と、ギフト・旅ログの基本構成を追加。",
        "effect": "最初のスマホモックを確認できる状態に。",
    },
    "Ignore local GitHub CLI files": {
        "detail": "ローカルの GitHub CLI 関連ファイルを git 管理対象から外した。",
        "effect": "ユーザー画面の変化なし。",
    },
    "Remove Pages workflow for initial push": {
        "detail": "GitHub Pages の公開方法を workflow からブランチ公開へ切り替えた。",
        "effect": "公開 URL を安定して使える状態に調整。",
    },
    "Convert mock to screen-based mobile flow": {
        "detail": "長い1ページ構成をやめて、画面遷移型のスマホモックへ変更。",
        "effect": "ポチポチ進めるアプリ風の体験に変更。",
    },
    "Add scenario confirmation step": {
        "detail": "旅候補を仮選択してから決定する、2段階の選択操作を追加。",
        "effect": "誤タップしづらい旅選択になった。",
    },
    "Compact mobile screens": {
        "detail": "余白、文字量、カードの高さを削って画面密度を調整。",
        "effect": "スクロールを減らしたコンパクト UI に改善。",
    },
    "Refine compact screen-first mock": {
        "detail": "iPhone アプリらしい見た目と操作導線になるよう再調整。",
        "effect": "よりスマホアプリっぽい印象に更新。",
    },
    "Refine ask-first mobile flow": {
        "detail": "おすすめ先行ではなく、相手・したい時間・要望入力から旅提案する流れに変更。",
        "effect": "相談してから旅が提案される体験に変わった。",
    },
    "Clarify journey recording flow": {
        "detail": "旅先ログをトグル記録にし、アルバムへつながる説明文を改善。",
        "effect": "押し間違いを戻しやすく、次の画面も理解しやすくなった。",
    },
    "Remove user-facing impact screen": {
        "detail": "地域価値ダッシュボードを導線から外し、アルバムで完了する流れに変更。",
        "effect": "最後がユーザー向けの完了画面として自然になった。",
    },
    "Add update history workbook": {
        "detail": "更新履歴の CSV / XLSX と、自動生成スクリプトを追加。",
        "effect": "更新内容を表形式で共有できるようになった。",
    },
    "Generate HTML update history page": {
        "detail": "更新履歴をカード型の HTML ページとして自動生成するようにした。",
        "effect": "スプレッドシートより見やすい履歴ページで確認できるようになった。",
    },
    "Refresh generated update history": {
        "detail": "最新の git 履歴をもとに、更新履歴ファイルを再生成した。",
        "effect": "履歴ページと履歴ファイルが最新状態になった。",
    },
}

HEADERS = [
    "更新日時",
    "コミット",
    "更新タイトル",
    "主な更新内容",
    "変更ファイル",
    "ユーザーに見える変化",
    "公開URL",
]


def record(full_hash: str, short_hash: str, updated_at: str, subject: str) -> dict[str, str]:
    detail = DETAIL_MAP.get(subject, {})
    return {
        "updated_at": updated_at,
        "full_hash": full_hash,
        "short_hash": short_hash,
        "subject": subject,
        "detail": detail.get("detail", subject),
        "files": changed_files(full_hash),
        "effect": detail.get("effect", "更新内容にあわせて反映。"),
        "public_url": PUBLIC_URL,
        "commit_url": f"{REPO_URL}/commit/{full_hash}",
    }


def run_git(*args: str) -> str:
    result = subprocess.run(
        ["git", "-C", str(ROOT), *args],
        check=True,
        capture_output=True,
        text=True,
        encoding="utf-8",
    )
    return result.stdout.strip()


def changed_files(commit_hash: str) -> str:
    raw = run_git("show", "--pretty=format:", "--name-only", "--no-renames", commit_hash)
    files = [line.strip() for line in raw.splitlines() if line.strip()]
    if not files:
        return "-"
    return ", ".join(files[:4])


def records() -> list[dict[str, str]]:
    raw = run_git(
        "log",
        "--date=format-local:%Y-%m-%d %H:%M",
        "--pretty=format:%H%x09%h%x09%ad%x09%s",
    )
    output: list[dict[str, str]] = []
    for line in raw.splitlines():
        full_hash, short_hash, updated_at, subject = line.split("\t", 3)
        output.append(record(full_hash, short_hash, updated_at, subject))
    return output


def rows(data: list[dict[str, str]]) -> list[list[str]]:
    return [
        [
            item["updated_at"],
            item["short_hash"],
            item["subject"],
            item["detail"],
            item["files"],
            item["effect"],
            item["public_url"],
        ]
        for item in data
    ]


def write_csv(data: list[dict[str, str]]) -> None:
    with CSV_PATH.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.writer(file)
        writer.writerow(HEADERS)
        writer.writerows(rows(data))


def write_xlsx(data: list[dict[str, str]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "更新履歴"
    sheet.append(HEADERS)
    for row in rows(data):
        sheet.append(row)

    header_fill = PatternFill(fill_type="solid", fgColor="1F1B16")
    header_font = Font(color="FFFFFF", bold=True)
    wrap = Alignment(vertical="top", wrap_text=True)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap

    widths = {
        1: 18,
        2: 12,
        3: 34,
        4: 42,
        5: 28,
        6: 36,
        7: 34,
    }
    for index, width in widths.items():
        sheet.column_dimensions[get_column_letter(index)].width = width

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:G{sheet.max_row}"

    note = workbook.create_sheet("運用メモ")
    note["A1"] = "更新方法"
    note["A2"] = "このファイルは tools/export_update_history.py を実行すると最新の git 履歴から再生成できます。"
    note["A3"] = "今後こちらでモックを更新したときも、このスクリプトを回して履歴を追記できます。"
    note["A1"].font = Font(bold=True)
    note.column_dimensions["A"].width = 90
    note["A2"].alignment = wrap
    note["A3"].alignment = wrap

    workbook.save(XLSX_PATH)


def stat_card(label: str, value: str) -> str:
    return f"""
      <article class="stat-card">
        <span>{html.escape(label)}</span>
        <strong>{html.escape(value)}</strong>
      </article>
    """


def file_tags(files: str) -> str:
    if files == "-":
        return '<span class="file-tag">-</span>'
    return "".join(
        f'<span class="file-tag">{html.escape(part.strip())}</span>'
        for part in files.split(",")
    )


def history_cards(data: list[dict[str, str]]) -> str:
    cards = []
    for index, item in enumerate(data):
        badge = "Latest" if index == 0 else f"#{len(data) - index}"
        cards.append(
            f"""
            <article class="history-card">
              <div class="card-top">
                <div>
                  <p class="eyebrow">{html.escape(item["updated_at"])}</p>
                  <h2>{html.escape(item["subject"])}</h2>
                </div>
                <span class="badge">{html.escape(badge)}</span>
              </div>
              <p class="detail">{html.escape(item["detail"])}</p>
              <div class="meta-grid">
                <section>
                  <p class="meta-label">ユーザーに見える変化</p>
                  <p class="meta-copy">{html.escape(item["effect"])}</p>
                </section>
                <section>
                  <p class="meta-label">変更ファイル</p>
                  <div class="file-tags">{file_tags(item["files"])}</div>
                </section>
              </div>
              <div class="link-row">
                <a href="{html.escape(item['commit_url'])}" target="_blank" rel="noreferrer">commit {html.escape(item["short_hash"])}</a>
                <a href="{html.escape(item['public_url'])}" target="_blank" rel="noreferrer">公開モック</a>
              </div>
            </article>
            """
        )
    return "".join(cards)


def write_html(data: list[dict[str, str]]) -> None:
    latest = data[0]["updated_at"] if data else "-"
    oldest = data[-1]["updated_at"] if data else "-"
    content = f"""<!DOCTYPE html>
<html lang="ja">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <meta name="description" content="旅録モックの更新履歴">
  <title>旅録 更新履歴</title>
  <style>
    :root {{
      --bg: #efe5d8;
      --bg-soft: #f7f1e8;
      --card: rgba(255, 253, 249, 0.84);
      --card-strong: #fffdf9;
      --ink: #191612;
      --muted: #71695e;
      --line: rgba(25, 22, 18, 0.08);
      --accent: #171717;
      --accent-soft: #ece2d4;
      --shadow: 0 24px 60px rgba(25, 22, 18, 0.08);
    }}

    * {{
      box-sizing: border-box;
    }}

    body {{
      margin: 0;
      font-family: "Zen Kaku Gothic New", "Hiragino Sans", sans-serif;
      color: var(--ink);
      background:
        radial-gradient(circle at top right, rgba(183, 156, 118, 0.22), transparent 28%),
        linear-gradient(180deg, #f3eadf 0%, #e4d7c5 100%);
    }}

    a {{
      color: inherit;
    }}

    .page {{
      width: min(1080px, calc(100% - 32px));
      margin: 0 auto;
      padding: 40px 0 56px;
    }}

    .hero {{
      display: grid;
      gap: 18px;
      padding: 28px;
      border: 1px solid var(--line);
      border-radius: 28px;
      background: linear-gradient(180deg, rgba(255, 255, 255, 0.7), rgba(247, 241, 232, 0.9));
      box-shadow: var(--shadow);
    }}

    .eyebrow {{
      margin: 0;
      color: var(--muted);
      font-size: 12px;
      letter-spacing: 0.14em;
      text-transform: uppercase;
    }}

    .hero h1 {{
      margin: 0;
      font-family: "Zen Old Mincho", serif;
      font-size: clamp(32px, 6vw, 56px);
      line-height: 1.08;
    }}

    .hero-copy {{
      max-width: 720px;
      margin: 0;
      color: var(--muted);
      font-size: 15px;
      line-height: 1.65;
    }}

    .hero-links {{
      display: flex;
      flex-wrap: wrap;
      gap: 10px;
    }}

    .hero-links a {{
      display: inline-flex;
      align-items: center;
      min-height: 42px;
      padding: 0 16px;
      border-radius: 999px;
      border: 1px solid var(--line);
      background: var(--card-strong);
      text-decoration: none;
    }}

    .hero-links a.primary {{
      background: var(--accent);
      border-color: var(--accent);
      color: #fff;
    }}

    .stats {{
      display: grid;
      grid-template-columns: repeat(3, minmax(0, 1fr));
      gap: 12px;
    }}

    .stat-card {{
      display: grid;
      gap: 4px;
      padding: 16px;
      border-radius: 22px;
      border: 1px solid var(--line);
      background: rgba(255, 255, 255, 0.56);
    }}

    .stat-card span {{
      color: var(--muted);
      font-size: 12px;
    }}

    .stat-card strong {{
      font-size: 26px;
      line-height: 1.1;
    }}

    .timeline {{
      position: relative;
      display: grid;
      gap: 14px;
      margin-top: 18px;
      padding-left: 22px;
    }}

    .timeline::before {{
      content: "";
      position: absolute;
      top: 8px;
      bottom: 8px;
      left: 7px;
      width: 2px;
      background: rgba(25, 22, 18, 0.12);
    }}

    .history-card {{
      position: relative;
      display: grid;
      gap: 14px;
      padding: 20px;
      border: 1px solid var(--line);
      border-radius: 24px;
      background: var(--card);
      box-shadow: var(--shadow);
      backdrop-filter: blur(10px);
    }}

    .history-card::before {{
      content: "";
      position: absolute;
      top: 26px;
      left: -21px;
      width: 14px;
      height: 14px;
      border-radius: 50%;
      background: var(--accent);
      box-shadow: 0 0 0 5px rgba(255, 248, 240, 0.95);
    }}

    .card-top {{
      display: flex;
      justify-content: space-between;
      gap: 16px;
      align-items: start;
    }}

    .card-top h2 {{
      margin: 4px 0 0;
      font-size: 24px;
      line-height: 1.2;
    }}

    .badge {{
      display: inline-flex;
      align-items: center;
      justify-content: center;
      min-height: 30px;
      padding: 0 12px;
      border-radius: 999px;
      background: var(--accent-soft);
      color: var(--ink);
      font-size: 12px;
      white-space: nowrap;
    }}

    .detail,
    .meta-copy {{
      margin: 0;
      color: var(--muted);
      line-height: 1.6;
    }}

    .meta-grid {{
      display: grid;
      grid-template-columns: 1.4fr 1fr;
      gap: 16px;
    }}

    .meta-label {{
      margin: 0 0 8px;
      color: var(--muted);
      font-size: 12px;
      letter-spacing: 0.08em;
      text-transform: uppercase;
    }}

    .file-tags {{
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
    }}

    .file-tag {{
      display: inline-flex;
      align-items: center;
      min-height: 28px;
      padding: 0 10px;
      border-radius: 999px;
      background: rgba(23, 23, 23, 0.06);
      font-size: 12px;
    }}

    .link-row {{
      display: flex;
      flex-wrap: wrap;
      gap: 12px;
    }}

    .link-row a {{
      display: inline-flex;
      align-items: center;
      min-height: 36px;
      padding: 0 14px;
      border-radius: 999px;
      border: 1px solid var(--line);
      background: rgba(255, 255, 255, 0.72);
      text-decoration: none;
      font-size: 13px;
    }}

    @media (max-width: 720px) {{
      .page {{
        width: min(100% - 20px, 100%);
        padding: 16px 0 32px;
      }}

      .hero {{
        padding: 18px;
        border-radius: 22px;
      }}

      .stats,
      .meta-grid {{
        grid-template-columns: 1fr;
      }}

      .timeline {{
        padding-left: 18px;
      }}

      .history-card {{
        padding: 16px;
        border-radius: 20px;
      }}

      .history-card::before {{
        left: -17px;
      }}

      .card-top {{
        display: grid;
      }}

      .card-top h2 {{
        font-size: 19px;
      }}
    }}
  </style>
  <link rel="preconnect" href="https://fonts.googleapis.com">
  <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
  <link href="https://fonts.googleapis.com/css2?family=Zen+Kaku+Gothic+New:wght@400;500;700&family=Zen+Old+Mincho:wght@500;700&display=swap" rel="stylesheet">
</head>
<body>
  <main class="page">
    <section class="hero">
      <p class="eyebrow">Tabiroku Update History</p>
      <h1>旅録モックの<br>更新履歴</h1>
      <p class="hero-copy">スプレッドシートより読みやすいように、更新内容をカード型のタイムラインにまとめたページです。git 履歴から自動生成しているので、今後の更新もこのページに増えていきます。</p>
      <div class="hero-links">
        <a class="primary" href="{html.escape(PUBLIC_URL)}" target="_blank" rel="noreferrer">公開モックを見る</a>
        <a href="{html.escape(REPO_URL)}" target="_blank" rel="noreferrer">GitHub Repo</a>
      </div>
      <div class="stats">
        {stat_card("更新数", str(len(data)))}
        {stat_card("最新更新", latest)}
        {stat_card("初回更新", oldest)}
      </div>
    </section>
    <section class="timeline">
      {history_cards(data)}
    </section>
  </main>
</body>
</html>
"""
    HTML_PATH.write_text(content, encoding="utf-8")


def main() -> None:
    data = records()
    write_csv(data)
    write_xlsx(data)
    write_html(data)
    print(f"created: {CSV_PATH}")
    print(f"created: {XLSX_PATH}")
    print(f"created: {HTML_PATH}")


if __name__ == "__main__":
    main()
